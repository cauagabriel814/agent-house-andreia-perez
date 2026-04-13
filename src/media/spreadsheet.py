import csv
import io

from src.services.uazapi import UazapiService
from src.utils.logger import logger

_MAX_ROWS = 20  # Limite de linhas exibidas por aba


async def process_spreadsheet(media_url: str, mimetype: str | None = None) -> str:
    """
    Le planilhas XLSX ou CSV recebidas via WhatsApp e retorna dados formatados.

    Pipeline:
        1. Download da planilha via UAZAPI
        2. Deteccao do tipo (XLSX ou CSV) pelo mimetype ou extensao da URL
        3. Parse dos dados com openpyxl (XLSX) ou csv (CSV)
        4. Retorna representacao textual estruturada (cabecalho + linhas)
    """
    if not media_url:
        return "[Planilha recebida sem URL de midia]"

    logger.info("SPREADSHEET | Iniciando download | url=%s", media_url)
    uazapi = UazapiService()
    data = await uazapi.download_media(media_url)

    if not data:
        return "[Erro: planilha vazia ou inacessivel]"

    logger.info("SPREADSHEET | Download concluido | bytes=%d", len(data))

    mime = (mimetype or "").lower()
    url_lower = media_url.lower()

    if "csv" in mime or url_lower.endswith(".csv"):
        result = _parse_csv(data)
    elif "xlsx" in mime or "spreadsheetml" in mime or url_lower.endswith(".xlsx"):
        result = _parse_xlsx(data)
    elif url_lower.endswith(".xls"):
        return "[Formato .xls antigo nao suportado. Por favor, envie em XLSX ou CSV.]"
    else:
        # Tentar XLSX primeiro, depois CSV
        result = _try_all(data)

    if not result:
        return "[Planilha sem dados extraiveis]"

    logger.info("SPREADSHEET | Parse concluido | chars=%d", len(result))
    return result


def _parse_csv(data: bytes) -> str:
    """Parseia um arquivo CSV e retorna representacao textual."""
    try:
        text = data.decode("utf-8")
    except UnicodeDecodeError:
        text = data.decode("latin-1")

    reader = csv.reader(text.splitlines())
    rows = list(reader)

    if not rows:
        return "[CSV vazio]"

    return _format_rows(rows, sheet_name=None)


def _parse_xlsx(data: bytes) -> str:
    """Parseia um arquivo XLSX e retorna representacao textual de todas as abas."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    parts = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(max_row=_MAX_ROWS + 1, values_only=True))
        total_rows = ws.max_row or 0

        formatted = _format_rows(
            [[str(c) if c is not None else "" for c in row] for row in rows],
            sheet_name=sheet_name,
            total_rows=total_rows,
        )
        if formatted:
            parts.append(formatted)

    wb.close()
    return "\n\n".join(parts)


def _format_rows(
    rows: list[list[str]],
    sheet_name: str | None,
    total_rows: int = 0,
) -> str:
    """Formata lista de linhas como texto legivel."""
    if not rows:
        return ""

    lines = []

    if sheet_name:
        lines.append(f"=== Aba: {sheet_name} ===")

    headers = [str(c) for c in rows[0]]
    separator = " | ".join(headers)
    lines.append(separator)
    lines.append("-" * min(len(separator), 80))

    for row in rows[1 : _MAX_ROWS + 1]:
        lines.append(" | ".join(str(c) for c in row))

    if total_rows > _MAX_ROWS + 1:
        lines.append(f"... e mais {total_rows - _MAX_ROWS - 1} linhas nao exibidas")

    return "\n".join(lines)


def _try_all(data: bytes) -> str:
    """Tenta parsear testando todos os formatos suportados."""
    for parser in (_parse_xlsx, _parse_csv):
        try:
            result = parser(data)
            if result.strip():
                return result
        except Exception:
            continue
    return ""
